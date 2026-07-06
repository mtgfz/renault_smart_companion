from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

NAVY = "1F2A5E"
ORANGE = "F5A623"
LIGHT = "EEF1F8"
GREEN = "2E7D32"
RED = "C62828"
WHITE = "FFFFFF"
FN = "Arial"

title_font = Font(name=FN, size=15, bold=True, color=WHITE)
subtitle_font = Font(name=FN, size=10, italic=True, color="555555")
header_font = Font(name=FN, size=10, bold=True, color=WHITE)
section_font = Font(name=FN, size=10, bold=True, color=WHITE)
normal_font = Font(name=FN, size=9, color="000000")
bold_font = Font(name=FN, size=9, bold=True, color="000000")
input_font = Font(name=FN, size=9, color="0000CC")
note_font = Font(name=FN, size=8, italic=True, color="777777")
adv_font = Font(name=FN, size=9, color=GREEN)
disadv_font = Font(name=FN, size=9, color=RED)

navy_fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
orange_fill = PatternFill("solid", start_color=ORANGE, end_color=ORANGE)
light_fill = PatternFill("solid", start_color=LIGHT, end_color=LIGHT)
white_fill = PatternFill("solid", start_color=WHITE, end_color=WHITE)

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = ["Option / Technologie", "Fournisseur / Type", "Coût mensuel (MAD)", "Coût unique (MAD)",
        "Coût estimé Année 1 (MAD)", "Avantages", "Inconvénients", "Recommandé si..."]
WIDTHS = [30, 22, 15, 15, 17, 42, 42, 34]


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def build_domain_sheet(name, intro, rows):
    ws = wb.create_sheet(name[:31])
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    ws["A1"] = name
    ws["A1"].font = title_font
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{get_column_letter(len(COLS))}2")
    ws["A2"] = intro
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 28

    header_row = 4
    for i, col in enumerate(COLS, start=1):
        ws.cell(row=header_row, column=i, value=col)
    style_header(ws, header_row, len(COLS))

    r = header_row + 1
    for option, provider, monthly, onetime, adv, disadv, reco in rows:
        ws.cell(row=r, column=1, value=option).font = bold_font
        ws.cell(row=r, column=2, value=provider).font = normal_font
        mc = ws.cell(row=r, column=3, value=monthly); mc.font = input_font; mc.number_format = "#,##0"
        oc = ws.cell(row=r, column=4, value=onetime); oc.font = input_font; oc.number_format = "#,##0"
        y1 = ws.cell(row=r, column=5, value=f"=C{r}*12+D{r}")
        y1.font = bold_font; y1.number_format = "#,##0"
        a = ws.cell(row=r, column=6, value=adv); a.font = adv_font; a.alignment = Alignment(wrap_text=True, vertical="top")
        d = ws.cell(row=r, column=7, value=disadv); d.font = disadv_font; d.alignment = Alignment(wrap_text=True, vertical="top")
        rc = ws.cell(row=r, column=8, value=reco); rc.font = note_font; rc.alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(1, len(COLS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c in (1, 2, 3, 4, 5):
                cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left" if c in (1, 2) else "right")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if (r - header_row) % 2 == 0:
                cell.fill = light_fill
        ws.row_dimensions[r].height = 46
        r += 1

    ws.freeze_panes = f"A{header_row + 1}"
    return ws


# =========================================================================
# 1. HÉBERGEMENT & SERVEURS
# =========================================================================
hebergement_rows = [
    ("Serveur local (on-premise / localhost)", "Matériel acheté par le concessionnaire",
     150, 18000,
     "Données 100% locales (confidentialité client renforcée, conformité CNDP simplifiée) ; aucun abonnement cloud lourd ; contrôle total du matériel.",
     "Pas d'accès à distance facile (sauf VPN à mettre en place) ; une panne matérielle arrête le service ; pas de scalabilité ; maintenance technique à charge du concessionnaire.",
     "Le concessionnaire veut garder les données en interne et dispose d'une petite salle serveur / d'un local technique."),
    ("VPS mutualisé", "OVH VPS Comfort / Hostinger Business",
     300, 0,
     "Très bon rapport qualité/prix ; accessible depuis n'importe où ; simple à administrer pour un seul concessionnaire.",
     "Ressources partagées avec d'autres clients ; pas de garantie de haute disponibilité ; capacité limitée si le trafic grandit fortement.",
     "Le volume de données/trafic reste celui d'un seul point de vente (cas le plus probable pendant et juste après le stage)."),
    ("Cloud managé international", "AWS (EC2/Lightsail) / Azure / Google Cloud",
     1800, 0,
     "Scalable à la demande ; très haute fiabilité (SLA élevés) ; sauvegardes automatiques ; écosystème riche (bases managées, Kubernetes...).",
     "Facturation à l'usage difficile à prévoir si mal configurée ; données hébergées hors Maroc (point d'attention CNDP) ; nécessite des compétences cloud.",
     "Le projet est validé pour un déploiement à plusieurs concessionnaires du réseau, pas seulement un site pilote."),
    ("Cloud régional / marocain", "Maroc Telecom Cloud / Inwi Business Cloud / datacenter local",
     1000, 0,
     "Données hébergées au Maroc (conformité CNDP facilitée) ; support en français/arabe ; latence réduite pour les utilisateurs locaux.",
     "Catalogue de services managés moins riche que les grands clouds internationaux ; moins d'automatisation avancée disponible.",
     "La conformité réglementaire locale (CNDP) et le support en langue locale sont prioritaires."),
    ("Kubernetes managé", "DigitalOcean DOKS / AWS EKS",
     2200, 0,
     "Orchestration automatique des services, haute disponibilité, bonne base pour un déploiement multi-sites futur.",
     "Coût et complexité largement injustifiés pour un seul concessionnaire ; nécessite une vraie équipe DevOps.",
     "Vision réseau national avec plusieurs concessionnaires connectés à la même plateforme (post-stage, pas pour le MVP)."),
]
build_domain_sheet(
    "1. Hébergement & Serveurs",
    "Où faire tourner l'application et les bases de données. Le choix dépend surtout du volume de données attendu et du niveau de confidentialité souhaité par le concessionnaire.",
    hebergement_rows,
)

# =========================================================================
# 2. FRONTEND (Web & Mobile)
# =========================================================================
frontend_rows = [
    ("Next.js (React) — Web", "Open source, développement interne/prestataire",
     0, 0,
     "Gratuit ; excellentes performances et SEO ; très large écosystème et communauté ; réutilisable pour d'autres projets Renault.",
     "Nécessite un développeur compétent en React/Next.js pour un résultat professionnel.",
     "Défaut recommandé pour l'app web (déjà le choix retenu dans l'architecture cible)."),
    ("React Native — Mobile", "Open source",
     0, 0,
     "Un seul code pour iOS et Android ; réutilise les compétences React de l'équipe web ; gratuit.",
     "Certaines fonctionnalités très spécifiques à l'OS nécessitent du code natif complémentaire.",
     "Le concessionnaire veut une app mobile sans doubler l'équipe de développement (iOS + Android séparés)."),
    ("Flutter — Mobile (alternative)", "Open source (Google)",
     0, 0,
     "Très bonnes performances et rendu visuel homogène ; un seul codebase pour iOS/Android.",
     "Langage Dart moins répandu au Maroc ; moins de développeurs disponibles sur le marché local.",
     "Une UI mobile très soignée est prioritaire et un développeur Flutter est disponible."),
    ("Solution low-code", "Bubble / Glide",
     600, 0,
     "Mise en place très rapide, sans développeur dédié à temps plein ; bon pour un premier prototype démo.",
     "Coût récurrent ; personnalisation limitée ; dépendance forte à la plateforme (vendor lock-in) ; difficile à faire évoluer vers l'architecture cible.",
     "Le concessionnaire veut un prototype visuel en quelques jours, sans engagement technique long terme."),
    ("Kit UI premium", "Tailwind UI (licence à vie) / shadcn/ui (gratuit)",
     0, 3000,
     "Gain de temps de design important ; composants déjà testés et accessibles.",
     "Tailwind UI est payant (coût unique) ; nécessite tout de même une adaptation à la charte graphique Renault.",
     "L'équipe veut un rendu visuel professionnel rapidement sans designer dédié (shadcn/ui suffit dans la plupart des cas, gratuit)."),
]
build_domain_sheet(
    "2. Frontend (Web & Mobile)",
    "Ce que voit et utilise le client final (configurateur, prise de RDV, chatbot). La plupart des briques recommandées sont open source et gratuites : le coût principal est le temps de développement, pas la licence.",
    frontend_rows,
)

# =========================================================================
# 3. BACKEND
# =========================================================================
backend_rows = [
    ("FastAPI (Python)", "Open source",
     0, 0,
     "Développement rapide ; excellent pour intégrer directement les modèles ML/IA ; bonnes performances asynchrones.",
     "Écosystème « entreprise » un peu moins mature que Java pour les très grandes organisations.",
     "Choix recommandé : l'équipe doit connecter le backend aux modèles de scoring et aux agents AI."),
    ("Spring Boot (Java)", "Open source",
     0, 0,
     "Très robuste et mature ; standard reconnu en entreprise ; bonne maintenabilité à long terme.",
     "Développement plus verbeux ; mise en place initiale plus longue que FastAPI.",
     "Le concessionnaire ou le groupe Renault impose déjà des standards Java en interne."),
    ("Node.js / Express", "Open source",
     0, 0,
     "Même langage que le frontend (JavaScript/TypeScript) ; bon pour les fonctionnalités temps réel (notifications, chat).",
     "Moins adapté aux traitements de Machine Learning lourds que Python.",
     "L'équipe est déjà majoritairement JavaScript et le volet ML reste externalisé/simple."),
    ("Firebase (Backend-as-a-Service)", "Google",
     400, 0,
     "Authentification, base de données et hébergement inclus ; mise en place très rapide sans DevOps.",
     "Coûts qui peuvent grimper vite avec le volume ; moins de contrôle sur l'emplacement des données.",
     "Besoin d'un prototype fonctionnel en quelques jours sans infrastructure à gérer."),
]
build_domain_sheet(
    "3. Backend",
    "La logique métier et les API qui connectent le frontend, les données et les modèles d'IA.",
    backend_rows,
)

# =========================================================================
# 4. BASE DE DONNÉES & STOCKAGE
# =========================================================================
data_rows = [
    ("PostgreSQL auto-hébergé", "Open source",
     0, 0,
     "Robuste, gratuit, excellent support des données relationnelles et JSON ; standard de l'industrie.",
     "Sauvegardes et mises à jour à la charge de l'équipe technique.",
     "Choix recommandé pour les données structurées (clients, ventes, leads)."),
    ("MySQL / MariaDB auto-hébergé", "Open source",
     0, 0,
     "Très répandu, énormément de documentation et de support communautaire.",
     "Moins riche que PostgreSQL pour l'analytique avancé (requêtes complexes, JSON).",
     "L'équipe est déjà familière de MySQL via d'autres projets internes."),
    ("Supabase (PostgreSQL managé)", "Supabase Cloud",
     235, 0,
     "Mise en place en quelques minutes ; authentification, API et stockage inclus ; sauvegardes automatiques.",
     "Dépendance à un fournisseur externe ; données hébergées hors Maroc.",
     "L'équipe veut aller vite sans gérer elle-même la base de données."),
    ("MongoDB Atlas (NoSQL managé)", "MongoDB Inc.",
     535, 0,
     "Flexible pour les données non structurées (logs de conversation, interactions agent AI).",
     "Moins adapté aux données relationnelles classiques (ventes, clients) ; coût récurrent.",
     "Beaucoup de données semi-structurées à stocker (historique des échanges avec les agents AI)."),
    ("Data lake — MinIO (auto-hébergé)", "Open source, compatible S3",
     0, 0,
     "Gratuit, compatible avec l'écosystème S3, contrôle total des données.",
     "Nécessite un serveur dédié avec suffisamment d'espace disque.",
     "Beaucoup de données brutes à stocker (logs, exports) sans dépendre d'un cloud externe."),
    ("Data lake — Amazon S3", "AWS",
     150, 0,
     "Fiable, scalable, facturé au volume réellement utilisé (~0,22 MAD/Go/mois).",
     "Facturé à l'usage, données hors Maroc.",
     "Le projet passe à l'échelle multi-concessionnaires et le volume de données devient important."),
]
build_domain_sheet(
    "4. Base de données & Stockage",
    "Où et comment sont stockées les données clients, ventes, leads et interactions.",
    data_rows,
)

# =========================================================================
# 5. STREAMING & DATA PIPELINE
# =========================================================================
streaming_rows = [
    ("File d'attente légère (Redis Streams / RabbitMQ)", "Open source, auto-hébergé",
     0, 0,
     "Largement suffisant pour le volume d'un seul concessionnaire ; simple à installer et à maintenir.",
     "Moins riche que Kafka pour du vrai temps réel à très grande échelle (non nécessaire ici).",
     "Choix recommandé pour le MVP et même au-delà, tant que le volume reste celui d'un seul site."),
    ("Apache Kafka auto-hébergé", "Open source, Docker",
     0, 0,
     "Standard de l'industrie ; très scalable ; directement réutilisable pour un futur déploiement multi-concessionnaires.",
     "Complexe à administrer sans une équipe DevOps dédiée ; largement surdimensionné pour 1 seul site.",
     "Le réseau Renault valide un déploiement à l'échelle de plusieurs concessionnaires en parallèle."),
    ("Confluent Cloud (Kafka managé)", "Confluent",
     2200, 0,
     "Aucune maintenance d'infrastructure à assurer ; scalabilité gérée automatiquement.",
     "Coût significatif pour un volume de données encore faible à l'échelle d'un concessionnaire.",
     "Le budget global permet d'externaliser complètement l'infrastructure data dès le départ."),
    ("Apache Airflow auto-hébergé", "Open source",
     250, 0,
     "Standard open source pour l'orchestration de pipelines (traitement quotidien des données) ; gratuit hormis la VM.",
     "Nécessite une petite VM dédiée et un minimum de maintenance.",
     "Dès que les traitements batch quotidiens (feature engineering, rapports) deviennent réguliers."),
    ("Astronomer (Airflow managé)", "Astronomer.io",
     2800, 0,
     "Zéro maintenance d'infrastructure, mises à jour automatiques.",
     "Coût élevé, nettement surdimensionné pour l'échelle d'un concessionnaire.",
     "Uniquement pertinent en cas de déploiement national avec de nombreux pipelines critiques."),
]
build_domain_sheet(
    "5. Streaming & Data Pipeline",
    "Comment les données circulent en temps réel entre les différents composants (site web, app, base de données, modèles).",
    streaming_rows,
)

# =========================================================================
# 6. MACHINE LEARNING
# =========================================================================
ml_rows = [
    ("Scikit-learn / XGBoost / LightGBM", "Open source, local ou petit serveur",
     0, 0,
     "Gratuit ; largement suffisant pour le scoring de propension et la segmentation à l'échelle d'un concessionnaire ; pas de dépendance cloud.",
     "Pas de GPU disponible, donc limité pour du deep learning lourd (non nécessaire pour ce cas d'usage).",
     "Choix recommandé : le volume de données d'un concessionnaire ne justifie pas plus."),
    ("Google Colab / Colab Pro", "Google",
     95, 0,
     "Accès GPU sans investissement matériel ; pratique pour le prototypage rapide pendant le stage.",
     "Sessions limitées dans le temps ; non adapté à un service de production permanent.",
     "Phase d'expérimentation et d'entraînement de modèles pendant les 2 mois de stage."),
    ("MLflow auto-hébergé", "Open source",
     0, 0,
     "Suivi des expériences et versionning des modèles ; standard open source largement utilisé.",
     "Nécessite une petite VM dédiée (déjà comptée dans l'hébergement).",
     "Dès qu'il y a plus d'un modèle en concurrence à comparer/versionner."),
    ("AWS Sagemaker / GCP Vertex AI", "Cloud managé",
     1200, 0,
     "MLOps complet clé en main (entraînement, déploiement, monitoring des modèles).",
     "Largement surdimensionné et coûteux pour le volume de données d'un seul concessionnaire.",
     "Déploiement à l'échelle de plusieurs dizaines de concessionnaires avec une équipe data dédiée."),
]
build_domain_sheet(
    "6. Machine Learning",
    "Modèles de scoring de propension à l'achat, segmentation clients, recommandation de véhicules/offres.",
    ml_rows,
)

# =========================================================================
# 7. REINFORCEMENT LEARNING
# =========================================================================
rl_rows = [
    ("Implémentation maison (bandits contextuels)", "Python / NumPy / scikit-learn",
     0, 0,
     "Gratuit ; suffisant pour personnaliser les offres à l'échelle d'un concessionnaire ; contrôle total de la logique.",
     "Nécessite une bonne maîtrise théorique des bandits contextuels pour une implémentation correcte.",
     "Choix recommandé pour le MVP et la première année d'exploitation."),
    ("Vowpal Wabbit", "Open source",
     0, 0,
     "Bibliothèque éprouvée spécifiquement conçue pour les bandits contextuels ; très performante et légère.",
     "Documentation moins accessible aux débutants que scikit-learn.",
     "L'équipe veut une solution de bandit contextuel plus mature et testée en production ailleurs."),
    ("Ray RLlib (Deep RL)", "Open source + calcul GPU cloud à l'usage",
     500, 0,
     "Permet d'évoluer vers du RL plus avancé si le volume de données grandit fortement.",
     "Complexité et coût GPU injustifiés tant que le volume de données reste celui d'un seul concessionnaire.",
     "Vision à 2-3 ans avec un volume de données conséquent sur plusieurs concessionnaires."),
]
build_domain_sheet(
    "7. Reinforcement Learning",
    "Personnalisation séquentielle des offres montrées à chaque client (quelle offre, à qui, quand).",
    rl_rows,
)

# =========================================================================
# 8. AGENTS AI / LLM
# =========================================================================
llm_rows = [
    ("Claude API (Anthropic)", "Cloud, facturation au token",
     2500, 0,
     "Excellent raisonnement ; très bon pour orchestrer des agents complexes multi-étapes (LangGraph, RAG).",
     "Coût proportionnel au volume de conversations ; les données transitent chez un tiers.",
     "Les agents doivent gérer des scénarios complexes (financement, SAV) nécessitant un raisonnement fin."),
    ("GPT API (OpenAI)", "Cloud, facturation au token",
     2500, 0,
     "Écosystème mature, bonne qualité générale, large documentation.",
     "Mêmes remarques que Claude : coût au volume, données envoyées à un tiers.",
     "Alternative à Claude selon les préférences de l'équipe ou les accords existants avec OpenAI."),
    ("Groq API", "Cloud, inférence rapide sur modèles open source",
     600, 0,
     "Latence très faible (idéal pour un chatbot showroom réactif) ; tarifs très compétitifs, tier gratuit généreux.",
     "Catalogue de modèles plus limité que Claude/GPT (moins adapté aux tâches de raisonnement très complexes).",
     "Priorité à la rapidité de réponse pour les interactions simples (FAQ, disponibilité véhicule, prise de RDV)."),
    ("LLM open source auto-hébergé (Ollama + Llama/Mistral)", "Open source, infra GPU dédiée",
     1500, 20000,
     "Données 100% locales (confidentialité maximale, conformité CNDP renforcée) ; pas de coût par requête une fois l'infra en place.",
     "Qualité générale un cran en dessous des meilleurs modèles propriétaires ; investissement infra GPU nécessaire.",
     "La confidentialité des échanges clients est une priorité absolue pour le concessionnaire/le groupe."),
]
build_domain_sheet(
    "8. Agents AI - LLM",
    "Le « moteur de conversation » des agents (véhicule, offres/financement, SAV). Plusieurs fournisseurs peuvent être combinés selon le cas d'usage.",
    llm_rows,
)

# =========================================================================
# 9. SÉCURITÉ
# =========================================================================
secu_rows = [
    ("Cloudflare Free", "Cloudflare",
     0, 0,
     "Protection DDoS de base et SSL gratuits ; mise en place en quelques minutes.",
     "Fonctionnalités de pare-feu applicatif (WAF) limitées.",
     "Point de départ minimal, suffisant pour un MVP peu exposé."),
    ("Cloudflare Pro", "Cloudflare",
     190, 0,
     "Pare-feu applicatif (WAF) plus complet, règles de sécurité avancées, meilleure protection.",
     "Coût récurrent, même si modéré.",
     "L'app gère des données sensibles (financement, coordonnées clients) et est exposée publiquement."),
    ("Let's Encrypt (certificat SSL)", "Open source / gratuit",
     0, 0,
     "Chiffrement HTTPS standard et automatisable, gratuit à vie.",
     "Renouvellement tous les 90 jours (facilement automatisable, donc impact quasi nul).",
     "Toujours recommandé, quel que soit le scénario retenu."),
    ("Audit sécurité / pentest externe (1x/an)", "Prestataire externe",
     0, 25000,
     "Identifie les failles avant un incident réel ; rassure la direction et les clients sur la fiabilité de la plateforme.",
     "Coût non négligeable, à renouveler chaque année.",
     "Avant toute mise en production réelle avec des données clients sensibles (recommandé dès la cible production)."),
    ("Conformité CNDP (déclaration traitement de données)", "Accompagnement juridique + frais de dossier",
     0, 3000,
     "Obligation légale au Maroc pour tout traitement de données personnelles ; évite les sanctions.",
     "Démarche administrative à anticiper (délais de traitement du dossier).",
     "Dès que des données personnelles de clients réels sont collectées et traitées (donc dès la cible production)."),
]
build_domain_sheet(
    "9. Sécurité",
    "Protection de l'application et conformité réglementaire (CNDP) pour le traitement des données clients.",
    secu_rows,
)

# =========================================================================
# 10. MLOPS & MONITORING
# =========================================================================
mlops_rows = [
    ("Prometheus + Grafana auto-hébergés", "Open source",
     0, 0,
     "Gratuit (hors VM), standard open source très personnalisable pour suivre les KPIs et la santé du système.",
     "Nécessite une VM dédiée et un peu de configuration initiale.",
     "Choix recommandé : partage la VM déjà utilisée pour l'hébergement (coût marginal quasi nul)."),
    ("Grafana Cloud (managé)", "Grafana Labs",
     460, 0,
     "Aucune maintenance d'infrastructure, mise en place très rapide.",
     "Coût récurrent qui augmente avec le volume de métriques suivies.",
     "L'équipe technique interne est réduite et ne veut gérer aucune infrastructure de monitoring."),
]
build_domain_sheet(
    "10. MLOps & Monitoring",
    "Suivi en temps réel de la santé de la plateforme et de la performance des modèles (dérive, erreurs, latence).",
    mlops_rows,
)

# =========================================================================
# GUIDE (cover sheet)
# =========================================================================
guide = wb.create_sheet("Guide")
guide.sheet_view.showGridLines = False
guide.column_dimensions["A"].width = 95

guide.merge_cells("A1:A1")
guide["A1"] = "Renault Smart Companion — Concessionnaire"
guide["A1"].font = Font(name=FN, size=18, bold=True, color=WHITE)
guide["A1"].fill = navy_fill
guide["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
guide.row_dimensions[1].height = 34

guide["A2"] = "Budget technologique détaillé — comparatif d'options par domaine"
guide["A2"].font = Font(name=FN, size=12, italic=True, color="555555")
guide.row_dimensions[2].height = 24

r = 4
guide.cell(row=r, column=1, value="Objectif de ce classeur").font = Font(name=FN, size=12, bold=True, color=NAVY); r += 1
guide.cell(row=r, column=1, value=(
    "Ce classeur ne propose pas un budget unique et figé : pour chaque brique technique du projet, plusieurs options "
    "réalistes sont comparées (coût mensuel, coût unique, avantages, inconvénients) afin que le concessionnaire "
    "puisse choisir librement le niveau d'investissement qui lui convient, domaine par domaine."
)).font = normal_font
guide.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
guide.row_dimensions[r].height = 60
r += 2

guide.cell(row=r, column=1, value="Hypothèses de calcul").font = Font(name=FN, size=12, bold=True, color=NAVY); r += 1
hyps = [
    "• 1 EUR ≈ 10,9 MAD ; 1 USD ≈ 9,4 MAD (taux indicatif juillet 2026, à ajuster selon le taux du jour).",
    "• « Coût unique » = investissement ponctuel (achat matériel, mise en place, licence à vie, audit annuel...).",
    "• « Coût estimé Année 1 » = (coût mensuel × 12) + coût unique, pour donner un ordre de grandeur sur la première année.",
    "• Toutes les estimations sont indicatives et à confirmer auprès des fournisseurs au moment de la décision.",
    "• Échelle de référence : un seul concessionnaire, pas un déploiement national.",
]
for h in hyps:
    guide.cell(row=r, column=1, value=h).font = normal_font
    r += 1
r += 1

guide.cell(row=r, column=1, value="Comment lire ce classeur").font = Font(name=FN, size=12, bold=True, color=NAVY); r += 1
sheets_desc = [
    "1. Hébergement & Serveurs — où faire tourner l'application (du serveur local au cloud managé).",
    "2. Frontend — technologies web/mobile visibles par le client.",
    "3. Backend — logique métier et API.",
    "4. Base de données & Stockage — où sont conservées les données.",
    "5. Streaming & Data Pipeline — comment les données circulent en temps réel.",
    "6. Machine Learning — scoring, segmentation, recommandation.",
    "7. Reinforcement Learning — personnalisation des offres.",
    "8. Agents AI / LLM — le moteur de conversation des agents.",
    "9. Sécurité — protection et conformité CNDP.",
    "10. MLOps & Monitoring — supervision de la plateforme.",
    "11. Synthèse & Scénarios — 3 combinaisons prêtes à présenter (Économique, Cloud managé, Hybride recommandé).",
]
for s in sheets_desc:
    guide.cell(row=r, column=1, value=s).font = normal_font
    r += 1

# =========================================================================
# SYNTHÈSE & SCÉNARIOS
# =========================================================================
synth = wb.create_sheet("11. Synthèse & Scénarios")
synth.sheet_view.showGridLines = False
for i, w in enumerate([26, 40, 18, 18], start=1):
    synth.column_dimensions[get_column_letter(i)].width = w

synth.merge_cells("A1:D1")
synth["A1"] = "Synthèse & Scénarios"
synth["A1"].font = title_font
synth["A1"].fill = navy_fill
synth["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
synth.row_dimensions[1].height = 26

synth.merge_cells("A2:D2")
synth["A2"] = "3 combinaisons possibles à partir des feuilles précédentes, pour aider la décision du concessionnaire."
synth["A2"].font = subtitle_font

scenarios = [
    {
        "name": "Scénario Économique (local / open source)",
        "color": "2E7D32",
        "choices": [
            ("Hébergement", "Serveur local (on-premise)"),
            ("Frontend", "Next.js + React Native (gratuit)"),
            ("Backend", "FastAPI (gratuit)"),
            ("Base de données", "PostgreSQL auto-hébergé (gratuit)"),
            ("Streaming", "Redis Streams auto-hébergé (gratuit)"),
            ("Machine Learning", "scikit-learn/XGBoost local (gratuit)"),
            ("Reinforcement Learning", "Implémentation maison (gratuit)"),
            ("Agents AI / LLM", "LLM open source auto-hébergé (Ollama)"),
            ("Sécurité", "Cloudflare Free + Let's Encrypt + CNDP"),
            ("MLOps", "Prometheus/Grafana auto-hébergés (gratuit)"),
        ],
        "monthly": 250,
        "onetime": 40000,
        "note": "Investissement de départ plus élevé (serveur + GPU), mais coût récurrent quasi nul ensuite. Demande une petite compétence technique en interne pour la maintenance.",
    },
    {
        "name": "Scénario Cloud managé complet",
        "color": "1565C0",
        "choices": [
            ("Hébergement", "Cloud managé international (AWS/Azure/GCP)"),
            ("Frontend", "Next.js + React Native (gratuit)"),
            ("Backend", "FastAPI sur le cloud (gratuit)"),
            ("Base de données", "Supabase managé"),
            ("Streaming", "Confluent Cloud (Kafka managé)"),
            ("Machine Learning", "AWS Sagemaker / GCP Vertex AI"),
            ("Reinforcement Learning", "Ray RLlib (GPU cloud à l'usage)"),
            ("Agents AI / LLM", "Claude API ou GPT API"),
            ("Sécurité", "Cloudflare Pro + pentest annuel + CNDP"),
            ("MLOps", "Grafana Cloud managé"),
        ],
        "monthly": 8885,
        "onetime": 3000,
        "note": "Zéro maintenance d'infrastructure interne, tout est géré par des prestataires. Coût récurrent élevé, largement surdimensionné pour un seul concessionnaire — pertinent surtout en vision réseau national.",
    },
    {
        "name": "Scénario Hybride recommandé",
        "color": "F5A623",
        "choices": [
            ("Hébergement", "VPS mutualisé (OVH/Hostinger)"),
            ("Frontend", "Next.js + React Native (gratuit)"),
            ("Backend", "FastAPI (gratuit)"),
            ("Base de données", "PostgreSQL auto-hébergé sur le VPS (gratuit)"),
            ("Streaming", "Redis Streams auto-hébergé (gratuit)"),
            ("Machine Learning", "scikit-learn local + Colab Pro ponctuel"),
            ("Reinforcement Learning", "Implémentation maison / Vowpal Wabbit (gratuit)"),
            ("Agents AI / LLM", "Groq API (rapide, économique) + Claude API en secours"),
            ("Sécurité", "Cloudflare Free + Let's Encrypt + pentest léger + CNDP"),
            ("MLOps", "Prometheus/Grafana auto-hébergés sur le VPS (gratuit)"),
        ],
        "monthly": 1245,
        "onetime": 18000,
        "note": "Le meilleur compromis pour un seul concessionnaire : coût maîtrisé, infrastructure simple à opérer, et une marge de manœuvre claire pour migrer vers plus de cloud managé si le volume de données augmente.",
    },
]

r = 4
for sc in scenarios:
    fill = PatternFill("solid", start_color=sc["color"], end_color=sc["color"])
    synth.merge_cells(f"A{r}:D{r}")
    c = synth.cell(row=r, column=1, value=sc["name"])
    c.font = Font(name=FN, size=12, bold=True, color=WHITE)
    c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    synth.row_dimensions[r].height = 22
    r += 1

    header_r = r
    for i, h in enumerate(["Domaine", "Choix retenu"], start=1):
        synth.cell(row=header_r, column=i, value=h).font = bold_font
        synth.cell(row=header_r, column=i).border = border
    r += 1
    for domain, choice in sc["choices"]:
        synth.cell(row=r, column=1, value=domain).font = normal_font
        synth.cell(row=r, column=2, value=choice).font = normal_font
        for c_idx in (1, 2):
            synth.cell(row=r, column=c_idx).border = border
            if (r - header_r) % 2 == 0:
                synth.cell(row=r, column=c_idx).fill = light_fill
        r += 1

    synth.cell(row=r, column=1, value="Coût mensuel récurrent estimé (MAD)").font = bold_font
    mcell = synth.cell(row=r, column=2, value=sc["monthly"]); mcell.font = bold_font; mcell.number_format = "#,##0"
    r += 1
    synth.cell(row=r, column=1, value="Coûts uniques estimés Année 1 (MAD)").font = bold_font
    ocell = synth.cell(row=r, column=2, value=sc["onetime"]); ocell.font = bold_font; ocell.number_format = "#,##0"
    r += 1
    synth.cell(row=r, column=1, value="TOTAL estimé Année 1 (MAD)").font = Font(name=FN, bold=True, size=11)
    total_cell = synth.cell(row=r, column=2, value=f"=B{r-2}*12+B{r-1}")
    total_cell.font = Font(name=FN, bold=True, size=11)
    total_cell.number_format = "#,##0"
    total_cell.fill = orange_fill
    r += 1
    synth.merge_cells(f"A{r}:D{r}")
    note_cell = synth.cell(row=r, column=1, value=sc["note"])
    note_cell.font = note_font
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    synth.row_dimensions[r].height = 44
    r += 2

# Reorder sheets: Guide first, domain sheets, Synthèse last, remove default empty sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

order = ["Guide", "1. Hébergement & Serveurs", "2. Frontend (Web & Mobile)", "3. Backend",
         "4. Base de données & Stockage", "5. Streaming & Data Pipeline",
         "6. Machine Learning", "7. Reinforcement Learning", "8. Agents AI - LLM",
         "9. Sécurité", "10. MLOps & Monitoring", "11. Synthèse & Scénarios"]
wb._sheets = [wb[name] for name in order]

wb.save("/home/claude/renault-smart-companion/docs/budget.xlsx")
print("saved")
