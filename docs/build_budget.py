from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

NAVY = "1F2A5E"
ORANGE = "F5A623"
LIGHT = "EEF1F8"
WHITE = "FFFFFF"

font_name = "Arial"
title_font = Font(name=font_name, size=16, bold=True, color=WHITE)
header_font = Font(name=font_name, size=11, bold=True, color=WHITE)
section_font = Font(name=font_name, size=11, bold=True, color=NAVY)
normal_font = Font(name=font_name, size=10, color="000000")
input_font = Font(name=font_name, size=10, color="0000FF")
note_font = Font(name=font_name, size=9, italic=True, color="777777")

navy_fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
orange_fill = PatternFill("solid", start_color=ORANGE, end_color=ORANGE)
light_fill = PatternFill("solid", start_color=LIGHT, end_color=LIGHT)

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, ncols, fill=navy_fill, font=header_font):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

# ---------------------------------------------------------------
# Sheet 1: Hypothèses
# ---------------------------------------------------------------
ws0 = wb.active
ws0.title = "Hypothèses"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 42
ws0.column_dimensions["B"].width = 22
ws0.column_dimensions["C"].width = 60

ws0.merge_cells("A1:C1")
ws0["A1"] = "Renault Smart Companion — Concessionnaire — Budget technique"
ws0["A1"].font = title_font
ws0["A1"].fill = navy_fill
ws0["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws0.row_dimensions[1].height = 30

ws0["A3"] = "Taux de change (référence)"
ws0["A3"].font = section_font
ws0["A4"], ws0["B4"] = "1 EUR =", 10.9
ws0["C4"] = "MAD — à ajuster selon le taux du jour"
ws0["A5"], ws0["B5"] = "1 USD =", 9.4
ws0["C5"] = "MAD — à ajuster selon le taux du jour"
for r in (4, 5):
    ws0[f"B{r}"].font = input_font
    ws0[f"B{r}"].number_format = "0.00"

ws0["A7"] = "Hypothèses du dossier d'investissement"
ws0["A7"].font = section_font
hyp = [
    ("Durée du stage (MVP)", "2 mois", "App web + dashboard + scoring ML de base"),
    ("Durée du financement demandé", "12 mois", "1 an d'exploitation post-stage, cible production"),
    ("Échelle", "1 concessionnaire", "Budget non mutualisé réseau — à diviser si déploiement multi-sites"),
    ("Hébergement", "Cloud (OVH/AWS, région EU/Maroc)", "Alternative on-premise possible, coût serveur physique en note"),
]
r = 8
for label, val, note in hyp:
    ws0.cell(row=r, column=1, value=label).font = normal_font
    c = ws0.cell(row=r, column=2, value=val); c.font = input_font
    ws0.cell(row=r, column=3, value=note).font = note_font
    r += 1

ws0["A14"] = "Comment lire ce classeur"
ws0["A14"].font = section_font
notes = [
    "• Onglet 'Budget MVP (2 mois)' : ce qui est réellement nécessaire pendant le stage — coût très faible, surtout du temps.",
    "• Onglet 'Budget Cible (12 mois)' : le dossier à présenter au concessionnaire pour financer la mise en production complète.",
    "• Onglet 'Synthèse' : totaux et graphique pour la présentation orale.",
    "• Les cellules en bleu sont modifiables (hypothèses) ; les totaux se recalculent automatiquement.",
]
r = 15
for n in notes:
    ws0.cell(row=r, column=1, value=n).font = note_font
    ws0.merge_cells(f"A{r}:C{r}")
    r += 1

# ---------------------------------------------------------------
# Sheet builder for budget sheets
# ---------------------------------------------------------------
COLS = ["Catégorie", "Poste", "Détail / Techno", "Type de coût", "Coût unitaire (MAD)", "Quantité / Durée", "Coût total (MAD)", "Source / Justification"]

def build_budget_sheet(title, rows, is_mvp=False):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    widths = [16, 24, 30, 16, 16, 14, 16, 42]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    header_row = 3
    for i, col in enumerate(COLS, start=1):
        ws.cell(row=header_row, column=i, value=col)
    style_header_row(ws, header_row, len(COLS))

    r = header_row + 1
    first_data_row = r
    current_section = None
    for row in rows:
        if row[0] == "SECTION":
            ws.merge_cells(f"A{r}:{get_column_letter(len(COLS))}{r}")
            cell = ws.cell(row=r, column=1, value=row[1])
            cell.font = section_font
            cell.fill = orange_fill
            cell.font = Font(name=font_name, size=11, bold=True, color=WHITE)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            r += 1
            continue
        cat, poste, detail, ctype, unit_cost, qty, source = row
        ws.cell(row=r, column=1, value=cat).font = normal_font
        ws.cell(row=r, column=2, value=poste).font = normal_font
        ws.cell(row=r, column=3, value=detail).font = normal_font
        ws.cell(row=r, column=4, value=ctype).font = normal_font
        cu = ws.cell(row=r, column=5, value=unit_cost); cu.font = input_font; cu.number_format = "#,##0"
        q = ws.cell(row=r, column=6, value=qty); q.font = input_font
        tot = ws.cell(row=r, column=7, value=f"=E{r}*F{r}")
        tot.font = normal_font
        tot.number_format = "#,##0"
        ws.cell(row=r, column=8, value=source).font = note_font
        for c in range(1, len(COLS) + 1):
            ws.cell(row=r, column=c).border = border
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = light_fill
        r += 1

    last_data_row = r - 1
    ws.cell(row=r, column=6, value="TOTAL (MAD)").font = Font(name=font_name, bold=True)
    ws.cell(row=r, column=6).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=r, column=7, value=f"=SUM(G{first_data_row}:G{last_data_row})")
    total_cell.font = Font(name=font_name, bold=True, size=11)
    total_cell.number_format = "#,##0"
    total_cell.fill = orange_fill
    ws.freeze_panes = f"A{header_row + 1}"
    return ws, first_data_row, last_data_row, r

# ---------------------------------------------------------------
# MVP budget (2 mois — stage)
# ---------------------------------------------------------------
mvp_rows = [
    ("SECTION", "Infrastructure & hébergement (MVP)"),
    ("Backend/Infra", "Serveur cloud dev (VPS)", "1 VPS 4 vCPU/8Go (OVH/AWS Lightsail) — dev + démo", "Récurrent mensuel", 350, 2, "Prix indicatif OVH VPS Comfort, 07/2026"),
    ("Backend/Infra", "Base de données PostgreSQL", "Managée ou incluse dans le VPS", "Récurrent mensuel", 0, 2, "Incluse dans le VPS pour le MVP"),
    ("Backend/Infra", "Nom de domaine + certificat SSL", "concessionnaire-renault-xxx.ma + Let's Encrypt (gratuit)", "Annuel", 150, 1, "Prix registrar .ma indicatif"),
    ("SECTION", "Outils & licences (MVP)"),
    ("Dev", "IDE / outils dev", "VS Code, GitHub gratuit (repo privé étudiant)", "Ponctuel", 0, 1, "Gratuit avec compte étudiant GitHub"),
    ("Agentic AI", "API LLM (agent conversationnel)", "Claude/GPT API — usage limité démo (~500k tokens)", "Ponctuel (période stage)", 400, 1, "Estimation tarif API, à confirmer selon fournisseur"),
    ("SECTION", "Sécurité (MVP)"),
    ("Sécurité", "Certificat SSL", "Let's Encrypt", "Gratuit", 0, 1, "Gratuit"),
    ("Sécurité", "Sauvegardes de base", "Snapshot hebdomadaire VPS", "Récurrent mensuel", 0, 2, "Inclus chez la plupart des hébergeurs VPS"),
    ("SECTION", "Ressources humaines"),
    ("RH", "Stagiaire PFA", "Développement, data, ML — temps plein 2 mois", "Forfait stage", 0, 1, "Selon politique de gratification stage de l'entreprise"),
]
ws1, first1, last1, total_row1 = build_budget_sheet("Budget MVP (2 mois)", mvp_rows, is_mvp=True)

# ---------------------------------------------------------------
# Target budget (12 months, production)
# ---------------------------------------------------------------
target_rows = [
    ("SECTION", "Frontend"),
    ("Frontend", "Développement app web (Next.js)", "Configurateur, RDV essai, agent conversationnel", "Ponctuel (dev externe/prestataire)", 25000, 1, "Estimation développement freelance/agence Maroc"),
    ("Frontend", "Développement app mobile (React Native)", "Version mobile de l'app utilisateur", "Ponctuel", 20000, 1, "Estimation développement freelance/agence Maroc"),
    ("Frontend", "Licence Tailwind UI / composants premium", "Kit UI professionnel (optionnel)", "Ponctuel", 3000, 1, "Prix catalogue Tailwind UI"),
    ("SECTION", "Agentic AI"),
    ("Agentic AI", "API LLM production (Claude/GPT)", "Volume estimé ~15M tokens/mois (trafic concessionnaire)", "Récurrent mensuel", 4500, 12, "Estimation tarifs API à confirmer selon fournisseur et volumétrie réelle"),
    ("Agentic AI", "Base vectorielle RAG (Qdrant Cloud)", "Plan starter/pro pour RAG catalogue", "Récurrent mensuel", 900, 12, "Tarif Qdrant Cloud plan starter, à vérifier"),
    ("SECTION", "Backend & Serving"),
    ("Backend", "Serveur applicatif (API FastAPI)", "1-2 vCPU dédiés, haute dispo basique", "Récurrent mensuel", 700, 12, "Estimation VPS/Cloud managé"),
    ("Backend", "OAuth2/Keycloak (auth)", "Auto-hébergé, coût = infra seule", "Récurrent mensuel", 0, 12, "Open source, pas de licence"),
    ("SECTION", "Streaming & traitement de données"),
    ("Data", "Cluster Kafka managé (léger)", "Confluent Cloud basic ou équivalent", "Récurrent mensuel", 1800, 12, "Estimation Confluent Cloud Basic, à ajuster selon volumétrie"),
    ("Data", "Orchestration Airflow (managé ou VM)", "1 VM dédiée ou Astronomer basic", "Récurrent mensuel", 900, 12, "Estimation, alternative auto-hébergée moins chère"),
    ("Data", "Traitement Spark/PySpark (à la demande)", "Jobs batch quotidiens, cluster éphémère", "Récurrent mensuel", 1200, 12, "Estimation cloud à l'usage (pay-as-you-go)"),
    ("SECTION", "Stockage"),
    ("Stockage", "Data lake (S3/MinIO)", "Stockage objet, volumétrie modérée", "Récurrent mensuel", 300, 12, "Estimation S3 standard, tarif au Go"),
    ("Stockage", "Entrepôt analytique (ClickHouse Cloud)", "Plan de base pour dashboards", "Récurrent mensuel", 1500, 12, "Estimation ClickHouse Cloud, à ajuster"),
    ("SECTION", "Machine Learning & MLOps"),
    ("ML", "Tracking MLflow (auto-hébergé)", "VM dédiée légère", "Récurrent mensuel", 250, 12, "Coût infra seule, MLflow open source"),
    ("ML", "Feature store (Feast, auto-hébergé)", "Coût infra seule", "Récurrent mensuel", 0, 12, "Open source"),
    ("MLOps", "Monitoring (Prometheus/Grafana)", "Auto-hébergé ou Grafana Cloud free/pro", "Récurrent mensuel", 400, 12, "Estimation Grafana Cloud Pro si managé"),
    ("MLOps", "Orchestration conteneurs (Kubernetes managé)", "1 cluster K8s managé (DOKS/EKS petit)", "Récurrent mensuel", 2200, 12, "Estimation cluster managé 3 nœuds, à ajuster selon charge"),
    ("SECTION", "Sécurité"),
    ("Sécurité", "Pare-feu applicatif (WAF) + DDoS", "Cloudflare Pro ou équivalent", "Récurrent mensuel", 250, 12, "Tarif Cloudflare Pro"),
    ("Sécurité", "Audit sécurité / pentest annuel", "Prestataire externe, 1x/an", "Annuel", 25000, 1, "Estimation prestation pentest PME au Maroc, variable selon périmètre"),
    ("Sécurité", "Conformité données (CNDP Maroc)", "Déclaration/autorisation traitement de données", "Ponctuel", 3000, 1, "Frais de dossier CNDP + accompagnement juridique, estimation"),
    ("Sécurité", "Certificats SSL, sauvegardes chiffrées", "Automatisées", "Récurrent mensuel", 150, 12, "Estimation"),
    ("SECTION", "Ressources humaines (12 mois post-stage)"),
    ("RH", "Data/ML engineer junior (temps partiel ou mi-temps)", "Maintenance, amélioration continue du modèle", "Récurrent mensuel", 8000, 12, "Estimation salaire junior mi-temps Maroc, à valider RH"),
    ("RH", "Formation équipe commerciale à l'outil", "1-2 sessions", "Ponctuel", 5000, 1, "Estimation prestation formation interne"),
]
ws2, first2, last2, total_row2 = build_budget_sheet("Budget Cible (12 mois)", target_rows)

# ---------------------------------------------------------------
# Synthèse sheet
# ---------------------------------------------------------------
ws3 = wb.create_sheet("Synthèse")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 34
ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 20

ws3.merge_cells("A1:C1")
ws3["A1"] = "Synthèse budgétaire"
ws3["A1"].font = title_font
ws3["A1"].fill = navy_fill
ws3["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws3.row_dimensions[1].height = 26

ws3["A3"] = "Scénario"
ws3["B3"] = "Coût total (MAD)"
ws3["C3"] = "Coût total (EUR, indicatif)"
style_header_row(ws3, 3, 3)

ws3["A4"] = "MVP (2 mois — stage)"
ws3["B4"] = f"='Budget MVP (2 mois)'!G{total_row1}"
ws3["C4"] = "=B4/Hypothèses!$B$4"
ws3["A5"] = "Cible (12 mois — production)"
ws3["B5"] = f"='Budget Cible (12 mois)'!G{total_row2}"
ws3["C5"] = "=B5/Hypothèses!$B$4"

for r in (4, 5):
    ws3[f"A{r}"].font = normal_font
    ws3[f"B{r}"].font = Font(name=font_name, bold=True)
    ws3[f"B{r}"].number_format = "#,##0"
    ws3[f"C{r}"].font = normal_font
    ws3[f"C{r}"].number_format = "#,##0"
    for c in ("A", "B", "C"):
        ws3[f"{c}{r}"].border = border

ws3["A7"] = "Note pour la présentation :"
ws3["A7"].font = section_font
ws3["A8"] = ("Le budget MVP correspond à ce qui est réellement engagé pendant les 2 mois de stage (très faible, "
             "principalement un VPS de développement). Le budget cible correspond au dossier d'investissement "
             "à soumettre à la direction du concessionnaire pour la mise en production sur 12 mois, avec un ROI "
             "attendu via l'augmentation du taux de conversion des leads (à chiffrer avec les données réelles "
             "du concessionnaire une fois collectées).")
ws3.merge_cells("A8:C8")
ws3["A8"].font = note_font
ws3["A8"].alignment = Alignment(wrap_text=True, vertical="top")
ws3.row_dimensions[8].height = 90

wb.save("/home/claude/renault-smart-companion/docs/budget.xlsx")
print("saved")
